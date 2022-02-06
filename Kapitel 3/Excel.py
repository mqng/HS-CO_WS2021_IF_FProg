import openpyxl
import math

wb = openpyxl.Workbook()

ws = wb.active
ws.title = "MeineErsteTabelle"

row = 1
for i in range(-31,32):
    x = i/10.0
    ws.cell(row, 1).value = x
    ws.cell(row, 2).value = x * math.sin(x)
    row += 1

wb.save("Datei16.xlsx")
print("Tabelle erstellt")
